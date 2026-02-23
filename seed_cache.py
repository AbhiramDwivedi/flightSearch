"""
seed_cache.py — Build .serp_cache.json from prior Excel exports.

Reads all flights_*.xlsx files in the project root, reconstructs minimal
SerpAPI-compatible response structures, then writes them into the local
response cache keyed against the CURRENT query's param sets.

This avoids re-spending SerpAPI credits for combinations already covered
by prior runs. Note: prior runs used include_airlines=F9 (Frontier only),
so the seeded cache will contain Frontier-only flights. Non-Frontier results
will be fetched fresh on the next run for any cache-miss combinations.

Usage:
    python seed_cache.py [--dry-run]
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT / "src"))

from flight_search import config          # noqa: E402
from flight_search.llm_parser import parse_query    # noqa: E402
from flight_search.flight_fetcher import _build_params, _cache_key  # noqa: E402

DRY_RUN = "--dry-run" in sys.argv

SEP = "→"   # unicode arrow used in route cells (U+2192)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _parse_duration_mins(s: str) -> int:
    """Convert '2h 46m' → 166."""
    m = re.match(r"(\d+)h\s*(\d+)m", str(s or ""))
    return int(m.group(1)) * 60 + int(m.group(2)) if m else 0


def _parse_route(route_str: str) -> tuple[str, str]:
    """'BWI→FLL' or 'BWI->FLL' → ('BWI', 'FLL')."""
    for sep in (SEP, "->", "-"):
        if sep in str(route_str):
            parts = str(route_str).split(sep, 1)
            return parts[0].strip(), parts[1].strip()
    return "", ""


def _make_segment(airline: str, depart_time: str, arrive_time: str,
                  orig: str, dest: str, tag: str) -> dict:
    """Build a minimal SerpAPI flight segment dict."""
    fn = f"{airline[:2].upper()}-SYN-{tag}"
    return {
        "airline": airline,
        "flight_number": fn,
        "airplane": "",
        "legroom": "",
        "travel_class": "Economy",
        "extensions": [],
        "departure_airport": {"id": orig, "time": str(depart_time)},
        "arrival_airport": {"id": dest, "time": str(arrive_time)},
    }


# ── Read all Excel files ───────────────────────────────────────────────────────

excel_files = sorted(ROOT.glob("flights_*.xlsx"), reverse=True)
if not excel_files:
    print("❌  No flights_*.xlsx files found in project root.")
    sys.exit(1)

print(f"📂  Found {len(excel_files)} Excel file(s):\n")
for f in excel_files:
    print(f"    {f.name}")
print()

# Structures to accumulate data across all files
# Keys: (outbound_date_str, return_date_str)
rt_groups:       dict[tuple, list[dict]] = {}  # Round-Trip groups (include return_flights)
ow_out_groups:   dict[tuple, list[dict]] = {}  # One-way outbound groups, keyed by (out_date, ret_date)
ow_ret_groups:   dict[tuple, list[dict]] = {}  # One-way return  groups, keyed by (out_date, ret_date)

# Dedup sets to avoid processing same flight twice across files
rt_seen:      set[str] = set()
ow_out_seen:  set[str] = set()
ow_ret_seen:  set[str] = set()

EXPECTED_HEADERS = [
    "Type", "Outbound Airline", "Outbound DateTime", "Outbound Arrive",
    "Outbound Route", "Outbound Duration", "Return Airline", "Return DateTime",
    "Return Arrive", "Return Route", "Return Duration",
    "Outbound Price", "Return Price", "Total Price",
]

for excel_path in excel_files:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [ws.cell(2, c).value for c in range(1, 15)]
    if headers != EXPECTED_HEADERS:
        print(f"  ⏭️   Skipping {excel_path.name} (incompatible format: {ws.max_column} cols)")
        continue
    print(f"  📄  Reading {excel_path.name}")

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            break  # End of data (empty row or footer)

        row_type     = str(row[0]  or "")
        o_airline    = str(row[1]  or "")
        o_depart     = str(row[2]  or "")   # "2026-03-29 21:43"
        o_arrive     = str(row[3]  or "")
        o_route      = str(row[4]  or "")   # "BWI→FLL"
        o_duration   = _parse_duration_mins(row[5])
        r_airline    = str(row[6]  or "")
        r_depart     = str(row[7]  or "")
        r_arrive     = str(row[8]  or "")
        r_route      = str(row[9]  or "")
        r_duration   = _parse_duration_mins(row[10])
        o_price      = int(row[11] or 0)
        r_price_raw  = row[12]
        r_price      = int(r_price_raw) if r_price_raw is not None else None

        o_orig, o_dest = _parse_route(o_route)
        r_orig, r_dest = _parse_route(r_route)

        o_date = o_depart[:10]  # "2026-03-29"
        r_date = r_depart[:10]  # "2026-04-06"
        date_key = (o_date, r_date)

        o_tag = f"{o_airline[:2].upper()}{o_depart.replace(' ','T').replace(':','')}"
        r_tag = f"{r_airline[:2].upper()}{r_depart.replace(' ','T').replace(':','')}"

        if "Round Trip" in row_type:
            dedup = f"RT|{o_tag}|{r_tag}"
            if dedup in rt_seen:
                continue
            rt_seen.add(dedup)

            o_seg = _make_segment(o_airline, o_depart, o_arrive, o_orig, o_dest, o_tag)
            r_seg = _make_segment(r_airline, r_depart, r_arrive, r_orig, r_dest, r_tag)

            group = {
                "flights": [o_seg],
                "layovers": [],
                "total_duration": o_duration,
                "price": o_price,
                "carbon_emissions": {},
                # Return leg attached (as _enrich_return_legs would do)
                "return_flights": [r_seg] if r_airline else [],
                "return_layovers": [],
                "return_total_duration": r_duration,
            }
            rt_groups.setdefault(date_key, []).append(group)

        elif "Independent One-Way" in row_type:
            # Outbound leg
            dedup_o = f"OW-OUT|{o_tag}"
            if dedup_o not in ow_out_seen:
                ow_out_seen.add(dedup_o)
                o_seg = _make_segment(o_airline, o_depart, o_arrive, o_orig, o_dest, o_tag)
                ow_out_groups.setdefault(date_key, []).append({
                    "flights": [o_seg],
                    "layovers": [],
                    "total_duration": o_duration,
                    "price": o_price,
                    "carbon_emissions": {},
                })

            # Return leg
            dedup_r = f"OW-RET|{r_tag}"
            if r_airline and dedup_r not in ow_ret_seen:
                ow_ret_seen.add(dedup_r)
                r_seg = _make_segment(r_airline, r_depart, r_arrive, r_orig, r_dest, r_tag)
                ow_ret_groups.setdefault(date_key, []).append({
                    "flights": [r_seg],
                    "layovers": [],
                    "total_duration": r_duration,
                    "price": r_price or 0,
                    "carbon_emissions": {},
                })

print(f"📊  Loaded from Excel:")
print(f"    Round-trip groups  : {sum(len(v) for v in rt_groups.values())} across {len(rt_groups)} date pair(s)")
print(f"    OW outbound groups : {sum(len(v) for v in ow_out_groups.values())} across {len(ow_out_groups)} date pair(s)")
print(f"    OW return groups   : {sum(len(v) for v in ow_ret_groups.values())} across {len(ow_ret_groups)} date pair(s)")
print()


# ── Parse current query to get new combination params ─────────────────────────

query_file = ROOT / "query.txt"
print("🤖  Parsing current query (reuses .last_parse.json if query unchanged)...")
q = query_file.read_text(encoding="utf-8")
parsed = parse_query(q)
print(f"    {len(parsed.combinations)} combinations found.\n")


# ── Build cache entries ────────────────────────────────────────────────────────

now_ts = datetime.now(timezone.utc).timestamp()
cache: dict = {}
seeded_rt = seeded_ow_out = seeded_ow_ret = 0

for combo in parsed.combinations:
    out_date = combo.outbound_date
    ret_date = combo.return_date or ""
    date_key = (out_date, ret_date)

    # ── Round-trip main call ───────────────────────────────────────────────────
    rt_data = rt_groups.get(date_key, [])
    if rt_data:
        params = _build_params(combo)
        key = _cache_key(params)
        cache[key] = {
            "timestamp": now_ts,
            "response": {
                "best_flights": rt_data,
                "other_flights": [],
                "__seeded_from_excel__": True,
            },
        }
        seeded_rt += 1
        print(f"  ✅  RT  [{out_date} → {ret_date}] seeded {len(rt_data)} group(s)")
    else:
        print(f"  ⬜  RT  [{out_date} → {ret_date}] no prior data — will fetch fresh")

    # ── Independent one-way outbound ───────────────────────────────────────────
    # Reproduce the exact param logic from fetch_all
    if combo.type == 1 and combo.return_date:
        base = _build_params(combo)
        base.pop("include_airlines", None)
        base.pop("exclude_airlines", None)

        # Outbound OW params
        out_p = dict(base)
        out_p["type"] = "2"
        out_p.pop("return_date", None)
        out_p.pop("return_times", None)
        if out_p.get("outbound_times"):
            parts = out_p["outbound_times"].split(",")
            if len(parts) == 4:
                out_p["outbound_times"] = f"{parts[0]},{parts[1]}"

        # Return OW params
        ret_p = dict(base)
        ret_p["type"] = "2"
        ret_p["departure_id"] = combo.arrival_id
        ret_p["arrival_id"] = combo.departure_id
        ret_p["outbound_date"] = combo.return_date
        ret_p.pop("return_date", None)
        ret_p.pop("outbound_times", None)
        if combo.return_times:
            rt_parts = combo.return_times.split(",")
            ret_p["outbound_times"] = (
                f"{rt_parts[0]},{rt_parts[1]}" if len(rt_parts) >= 2 else combo.return_times
            )

        ow_out_data = ow_out_groups.get(date_key, [])
        if ow_out_data:
            key = _cache_key(out_p)
            cache[key] = {
                "timestamp": now_ts,
                "response": {
                    "best_flights": ow_out_data,
                    "other_flights": [],
                    "__seeded_from_excel__": True,
                },
            }
            seeded_ow_out += 1
            print(f"  ✅  OW-out [{out_date}] seeded {len(ow_out_data)} group(s)")

        ow_ret_data = ow_ret_groups.get(date_key, [])
        if ow_ret_data:
            key = _cache_key(ret_p)
            cache[key] = {
                "timestamp": now_ts,
                "response": {
                    "best_flights": ow_ret_data,
                    "other_flights": [],
                    "__seeded_from_excel__": True,
                },
            }
            seeded_ow_ret += 1
            print(f"  ✅  OW-ret [{ret_date}] seeded {len(ow_ret_data)} group(s)")

print()
total_keys = len(cache)
print(f"📦  Cache entries prepared : {total_keys}")
print(f"    Round-trip seeded      : {seeded_rt}/{len(parsed.combinations)}")
print(f"    OW outbound seeded     : {seeded_ow_out}/{len(parsed.combinations)}")
print(f"    OW return seeded       : {seeded_ow_ret}/{len(parsed.combinations)}")
print(f"    TTL                    : {config.SERPAPI_CACHE_TTL_HOURS}h from now")
print()

if DRY_RUN:
    print("🔍  Dry run — cache NOT written.")
else:
    # Merge with any existing cache (don't overwrite unrelated entries)
    existing: dict = {}
    if config.CACHE_FILE.exists():
        try:
            existing = json.loads(config.CACHE_FILE.read_text(encoding="utf-8"))
            print(f"    Merging with {len(existing)} existing cache entries...")
        except (json.JSONDecodeError, OSError):
            pass
    existing.update(cache)
    config.CACHE_FILE.write_text(json.dumps(existing), encoding="utf-8")
    print(f"✅  Written to: {config.CACHE_FILE}")
    print()
    print("    Run the tool normally — seeded combinations will show '🗄️ from cache'.")
    print("    Combinations without prior data will fetch fresh from SerpAPI.")
