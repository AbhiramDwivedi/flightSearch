"""Excel exporter: writes FlightResult list to a formatted .xlsx file."""

from __future__ import annotations
import os
import statistics
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from .models import FlightResult


# ── Colour palette ─────────────────────────────────────────────────────────────
_HEADER_BG   = "2F5496"   # dark blue
_HEADER_FG   = "FFFFFF"   # white
_ROW_ALT     = "EEF2F9"   # light blue-grey (alternating rows)
_GREEN_FILL  = "C6EFCE"   # green  — price below median
_RED_FILL    = "FFC7CE"   # red    — price above 1.5 × median
_BORDER_CLR  = "BFBFBF"

# Column definitions: (header label, attribute on FlightResult, width)
_COLUMNS: list[tuple[str, str, int]] = [
    ("Preferred",         "preferred",                 11),
    ("Type",              "itinerary_type",            22),
    ("Outbound Airline",  "airline",                   20),
    ("Outbound DateTime", "depart_time",               22),
    ("Outbound Arrive",   "arrive_time",               22),
    ("Outbound Route",    "origin",                    16),
    ("Outbound Duration", "total_duration_mins",       16),
    ("Return Airline",    "return_airline",            20),
    ("Return DateTime",   "return_depart_time",        22),
    ("Return Arrive",     "return_arrive_time",        22),
    ("Return Route",      "destination",               16),
    ("Return Duration",   "return_total_duration_mins",16),
    ("Outbound Price",    "outbound_price",            14),
    ("Return Price",      "return_price",              14),
    ("Total Price",       "total_price",               14),
]


def _thin_border() -> Border:
    side = Side(style="thin", color=_BORDER_CLR)
    return Border(left=side, right=side, top=side, bottom=side)


def _fmt_duration(minutes: int) -> str:
    h, m = divmod(minutes, 60)
    return f"{h}h {m:02d}m"


def _route_value(result: FlightResult, is_return: bool = False) -> str:
    if is_return:
        return f"{result.destination}→{result.origin}" if result.return_depart_time else ""
    return f"{result.origin}→{result.destination}"


def export(results: list[FlightResult], query_summary: str, output_dir: Path = Path(".")) -> Path:
    """
    Write results to an xlsx file and return the file path.
    Auto-opens the file on Windows.
    """
    wb = Workbook()
    # wb.worksheets[0] is always Worksheet (not Optional) on a fresh Workbook
    ws: Worksheet = wb.worksheets[0]
    ws.title = "Flights"

    # ── Freeze pane & sheet title row ──────────────────────────────────────────
    title_text = f"Flight Results — {query_summary}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(name="Calibri", bold=True, size=13, color=_HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Header row ─────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F3864")
    for col_idx, (label, _, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(name="Calibri", bold=True, size=10, color=_HEADER_FG)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 28

    # ── Data rows ──────────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor=_ROW_ALT)
    for row_idx, result in enumerate(results, start=3):
        is_alt = (row_idx % 2 == 0)
        for col_idx, (_, attr, _) in enumerate(_COLUMNS, start=1):
            if attr == "origin":
                value = _route_value(result, is_return=False)
            elif attr == "destination":
                value = _route_value(result, is_return=True)
            else:
                value = getattr(result, attr)
            # Format duration column
            if attr in ("total_duration_mins", "return_total_duration_mins") and isinstance(value, int):
                value = _fmt_duration(value)
            if value is None:
                value = ""
            if attr == "preferred":
                value = "✓" if value else ""
            if attr == "itinerary_type":
                value = "Independent One-Way" if value == "independent_one_way" else "Round Trip"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (2, 5, 7, 10)))
            cell.border = _thin_border()
            if is_alt:
                cell.fill = alt_fill

    last_data_row = 2 + len(results)

    # ── Auto-filter (on header row) ────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:{get_column_letter(len(_COLUMNS))}{last_data_row}"

    # ── Freeze top 2 rows ──────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Conditional formatting on Total Price column (col 15 = "O") ─────────────────
    price_col = "O"
    price_range = f"{price_col}3:{price_col}{last_data_row}"
    prices = [
        (r.total_price if r.total_price is not None else r.price)
        for r in results
        if (r.total_price if r.total_price is not None else r.price) > 0
    ]
    median_price: float = statistics.median(prices) if prices else 0.0
    if prices:
        threshold_low  = int(median_price)
        threshold_high = int(median_price * 1.5)

        ws.conditional_formatting.add(
            price_range,
            CellIsRule(
                operator="lessThanOrEqual",
                formula=[str(threshold_low)],
                fill=PatternFill("solid", fgColor=_GREEN_FILL),
            ),
        )
        ws.conditional_formatting.add(
            price_range,
            CellIsRule(
                operator="greaterThan",
                formula=[str(threshold_high)],
                fill=PatternFill("solid", fgColor=_RED_FILL),
            ),
        )

    # ── Summary info at the bottom ────────────────────────────────────────────
    summary_row = last_data_row + 2
    ws.cell(row=summary_row, column=1, value=f"Total flights shown: {len(results)}")
    ws.cell(row=summary_row, column=1).font = Font(italic=True, size=9, color="808080")
    if prices:
        ws.cell(row=summary_row + 1, column=1,
                value=f"Price range: ${min(prices)} – ${max(prices)}  |  Median: ${int(median_price)}")
        ws.cell(row=summary_row + 1, column=1).font = Font(italic=True, size=9, color="808080")
        ws.cell(row=summary_row + 2, column=1,
                value="🟢 Green = at or below median price   🔴 Red = above 1.5× median price")
        ws.cell(row=summary_row + 2, column=1).font = Font(italic=True, size=9, color="808080")

    # ── Save ───────────────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = output_dir / f"flights_{timestamp}.xlsx"
    wb.save(filename)

    # Auto-open on Windows
    try:
        os.startfile(str(filename))
    except Exception:
        pass  # Non-Windows or no default app

    return filename
